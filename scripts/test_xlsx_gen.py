from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Presupuesto Mensual"

# Encabezados
ws['A1'] = "Concepto"
ws['B1'] = "Monto Estimado"
ws['C1'] = "Monto Real"
ws['D1'] = "Diferencia"

# Estilo para encabezados
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
for cell in ['A1', 'B1', 'C1', 'D1']:
    ws[cell].font = header_font
    ws[cell].fill = header_fill

# Datos (Inputs en AZUL según estándar)
blue_font = Font(color="0000FF") # RGB: 0,0,255
ws['A2'] = "Vivienda"
ws['B2'] = 1200
ws['B2'].font = blue_font
ws['C2'] = 1200
ws['C2'].font = blue_font

ws['A3'] = "Alimentación"
ws['B3'] = 400
ws['B3'].font = blue_font
ws['C3'] = 450
ws['C3'].font = blue_font

ws['A4'] = "Transporte"
ws['B4'] = 150
ws['B4'].font = blue_font
ws['C4'] = 130
ws['C4'].font = blue_font

# Totales y Fórmulas (En NEGRO por defecto)
ws['A6'] = "TOTAL"
ws['A6'].font = Font(bold=True)
ws['B6'] = "=SUM(B2:B4)"
ws['C6'] = "=SUM(C2:C4)"

# Cálculo de Diferencia (Fórmulas)
for row in range(2, 5):
    ws[f'D{row}'] = f"=C{row}-B{row}"

ws['D6'] = "=SUM(D2:D4)"

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15

wb.save("workspaces/personal/projects/presupuesto_test.xlsx")
print("Archivo 'presupuesto_test.xlsx' generado con éxito.")
